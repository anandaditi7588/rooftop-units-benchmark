"""
Generates the default Physical_Data.xlsx template, in the project root, if
it doesn't already exist. This is the "master template" referenced in the
spec: Column B holds the engineering parameters to benchmark. Column A is an
optional category and Column C an optional expected unit (both purely for
readability — only Column B is required by the extraction pipeline).

IMPORTANT: this never overwrites a Physical_Data.xlsx that's already there.
If you already have a real master template in the project folder, it is
used as-is and this generator never runs.

Run standalone with:  python -m scripts.generate_sample_physical_data
It is also called automatically on server startup if the file is missing.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

# (Category, Parameter, Expected Unit) — representative of a real commercial
# heat pump rooftop unit spec sheet. Engineers can freely edit this sheet or
# upload their own via the web UI ("Upload your own benchmark parameter Excel?").
DEFAULT_ROWS: list[tuple[str, str, str]] = [
    ("Capacity", "Cooling Capacity", "Btuh / Tons"),
    ("Capacity", "Heating Capacity", "Btuh"),
    ("Capacity", "Capacity Range", "Tons"),
    ("Efficiency", "COP", "-"),
    ("Efficiency", "EER", "Btuh/W"),
    ("Efficiency", "SEER", "-"),
    ("Efficiency", "IEER", "-"),
    ("Efficiency", "Thermal Efficiency", "%"),
    ("Physical", "Weight", "lbs"),
    ("Physical", "Dimensions", "in (L x W x H)"),
    ("Refrigeration", "Compressor Type", "-"),
    ("Refrigeration", "Compressor Stages", "steps"),
    ("Refrigeration", "Refrigerant", "-"),
    ("Refrigeration", "Refrigerant Circuits", "-"),
    ("Refrigeration", "Condenser Coil", "-"),
    ("Refrigeration", "Evaporator Coil", "-"),
    ("Airflow", "Supply Airflow", "CFM"),
    ("Airflow", "Static Pressure", "in. w.g."),
    ("Airflow", "Fan Motor", "HP"),
    ("Acoustics", "Sound Level", "dBA"),
    ("Electrical", "Electrical Requirements", "V/Ph/Hz"),
    ("Controls", "Economizer", "-"),
    ("Controls", "Defrost Control", "-"),
    ("Controls", "Controls", "-"),
    ("Filtration", "Filter Type", "MERV"),
    ("Operating Range", "Ambient Operating Range", "°F"),
    ("Heating", "Gas Heat Input", "MBH"),
    ("Heating", "Gas Heat Output", "MBH"),
    ("General", "Unit Type", "-"),
    ("General", "Warranty", "years"),
]


def ensure_sample_parameter_file(path: Path) -> Path:
    """Create Physical_Data.xlsx with sensible defaults if missing.
    Never overwrites an existing file — this is the user's master template."""
    if path.exists():
        return path

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Physical_Data"

    headers = ["Category", "Parameter", "Expected Unit"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F3B57")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for category, parameter, unit in DEFAULT_ROWS:
        ws.append([category, parameter, unit])

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18

    wb.save(path)
    return path


if __name__ == "__main__":
    from core.config import DEFAULT_PARAMETER_FILE

    out = ensure_sample_parameter_file(DEFAULT_PARAMETER_FILE)
    print(f"Physical_Data.xlsx ready at: {out}")
