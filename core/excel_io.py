"""
Excel I/O: reading the benchmark parameter template and writing the
formatted comparison workbook.

Reading side
------------
``read_parameters_from_excel`` reads Column B of the given workbook (per the
spec). Column A (if present) is treated as an optional category, Column C
(if present) as an optional unit. The first row is treated as a header and
skipped only if it doesn't look like a real parameter (heuristic: matches
common header words like "parameter").

Writing side
------------
``write_comparison_excel`` produces a professional, engineer-friendly
workbook:
  * Sheet "Comparison": Parameter x Competitor grid with the exact columns
    requested by the spec (Parameter, <competitor columns>, Source Document,
    Page Number, Confidence Score), aggregated across competitors for those
    last three columns.
  * Sheet "Sources_Detail": one row per (parameter, competitor) pair with
    the individual source document / page / confidence — this is where the
    full traceability lives.
  * Conditional formatting: missing values highlighted red, cross-competitor
    discrepancies highlighted yellow, best-in-class numeric values
    highlighted green (per config/parameter_rules.json directionality).
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from core.config import Competitor
from core.schemas import ParameterRow

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F3B57")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F6FA")
MISSING_FILL = PatternFill("solid", fgColor="F8CBCB")
DISCREPANCY_FILL = PatternFill("solid", fgColor="FFF3B0")
BEST_FILL = PatternFill("solid", fgColor="C6EFCE")
BEST_FONT = Font(bold=True, color="1E6B34")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
BANNER_FILL = PatternFill("solid", fgColor="0B2540")
BANNER_FONT = Font(bold=True, italic=True, color="7EE8C0", size=11)


def _write_query_banner(ws, text: str, num_columns: int) -> None:
    """A merged banner row at the very top of a sheet naming the specific
    series/model/unit-configuration this comparison was scoped to, so the
    Excel is self-explanatory even once it's been forwarded around."""
    last_col_letter = get_column_letter(num_columns)
    ws.merge_cells(f"A1:{last_col_letter}1")
    cell = ws.cell(row=1, column=1, value=f"Benchmark scope: {text}")
    cell.font = BANNER_FONT
    cell.fill = BANNER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

_HEADER_WORDS = {
    "parameter", "parameters", "engineering parameter", "item", "items",
    "spec", "specification", "field", "name",
}


def read_parameters_from_excel(path: Path) -> list[dict]:
    """Read benchmark parameters from Column B (A=category, C=unit if present).

    Column A commonly uses a merged cell spanning several related parameter
    rows (e.g. one "Compressor Data" category header merged over its
    Quantity/Type/Capacity Steps/Circuits rows) — openpyxl reads every cell
    in a merge except the top-left one back as None, so the category is
    forward-filled here from the last non-blank Column A cell seen. Without
    this, generic repeated parameter names like "Type" would lose the very
    context (which category they belong to) that makes them meaningful,
    and the matching engine would have no way to tell them apart.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows: list[dict] = []
    header_skipped = False
    current_category: Optional[str] = None

    for row in ws.iter_rows(min_row=1):
        col_a = row[0].value if len(row) > 0 else None
        col_b = row[1].value if len(row) > 1 else None
        col_c = row[2].value if len(row) > 2 else None

        if col_a is not None and str(col_a).strip():
            current_category = str(col_a).strip()

        if col_b is None or str(col_b).strip() == "":
            continue

        param_text = str(col_b).strip()

        if not header_skipped and param_text.lower() in _HEADER_WORDS:
            header_skipped = True
            # A header row's Column A ("Category") is a column label, not a
            # real category value — don't let it leak into forward-fill.
            current_category = None
            continue
        header_skipped = True

        rows.append({
            "category": current_category,
            "parameter": param_text,
            "unit": str(col_c).strip() if col_c else None,
        })

    logger.info("Loaded %d parameters from %s", len(rows), path)
    return rows


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

def write_comparison_excel(
    output_path: Path,
    parameter_rows: list[ParameterRow],
    competitors: list[Competitor],
    query: Optional[str] = None,
) -> None:
    wb = openpyxl.Workbook()

    _write_comparison_sheet(wb.active, parameter_rows, competitors, query)
    detail_ws = wb.create_sheet("Sources_Detail")
    _write_detail_sheet(detail_ws, parameter_rows, competitors, query)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Comparison workbook written to %s", output_path)


def _write_comparison_sheet(
    ws, parameter_rows: list[ParameterRow], competitors: list[Competitor], query: Optional[str] = None
) -> None:
    ws.title = "Comparison"

    headers = ["Category", "Parameter", "Unit"] + [c.name for c in competitors] + [
        "Source Document", "Page Number", "Confidence Score"
    ]

    offset = 0
    if query:
        _write_query_banner(ws, query, len(headers))
        offset = 1

    header_row = 1 + offset
    ws.append(headers)  # openpyxl appends after the current max_row (1 if a banner was just written)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Tint competitor columns with that competitor's brand color for quick scanning.
        comp_index = col_idx - 4  # 0-based index into competitors, once past Category/Parameter/Unit
        if 0 <= comp_index < len(competitors):
            cell.fill = PatternFill("solid", fgColor=competitors[comp_index].color.lstrip("#"))
        else:
            cell.fill = HEADER_FILL

    for r_idx, prow in enumerate(parameter_rows, start=header_row + 1):
        ws.cell(row=r_idx, column=1, value=prow.category or "")
        ws.cell(row=r_idx, column=2, value=prow.parameter)
        ws.cell(row=r_idx, column=3, value=prow.unit or "")

        sources, pages, confidences = [], [], []
        values_for_row: dict[str, Optional[str]] = {}

        for c_idx, comp in enumerate(competitors, start=4):
            cell_data = prow.values.get(comp.id)
            value = cell_data.value if cell_data else None
            cell = ws.cell(row=r_idx, column=c_idx, value=value or "")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            values_for_row[comp.id] = value

            if not value:
                cell.fill = MISSING_FILL
            elif prow.is_best_highlight.get(comp.id):
                cell.fill = BEST_FILL
                cell.font = BEST_FONT
            elif prow.has_discrepancy:
                cell.fill = DISCREPANCY_FILL

            if cell_data:
                if cell_data.source_document:
                    sources.append(cell_data.source_document)
                if cell_data.page_number:
                    pages.append(str(cell_data.page_number))
                if value:
                    confidences.append(cell_data.confidence)

        source_col = 4 + len(competitors)
        ws.cell(row=r_idx, column=source_col, value="; ".join(sorted(set(sources))))
        ws.cell(row=r_idx, column=source_col + 1, value="; ".join(sorted(set(pages))))
        avg_conf = round(sum(confidences) / len(confidences), 2) if confidences else 0.0
        ws.cell(row=r_idx, column=source_col + 2, value=avg_conf)

        if r_idx % 2 == 0:
            for col_idx in (1, 2, 3, source_col, source_col + 1, source_col + 2):
                fill = ws.cell(row=r_idx, column=col_idx).fill
                if fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(row=r_idx, column=col_idx).fill = ALT_ROW_FILL

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = THIN_BORDER

    ws.freeze_panes = f"D{header_row + 1}"
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{ws.max_row}"

    _autosize_columns(ws)


def _write_detail_sheet(
    ws, parameter_rows: list[ParameterRow], competitors: list[Competitor], query: Optional[str] = None
) -> None:
    headers = ["Parameter", "Competitor", "Value", "Source Document", "Page Number",
               "Confidence Score", "Matched Phrase"]

    offset = 0
    if query:
        _write_query_banner(ws, query, len(headers))
        offset = 1

    header_row = 1 + offset
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    r_idx = header_row + 1
    for prow in parameter_rows:
        for comp in competitors:
            cell_data = prow.values.get(comp.id)
            if cell_data is None:
                continue
            ws.cell(row=r_idx, column=1, value=prow.parameter)
            ws.cell(row=r_idx, column=2, value=comp.name)
            ws.cell(row=r_idx, column=3, value=cell_data.value or "")
            ws.cell(row=r_idx, column=4, value=cell_data.source_document or "")
            ws.cell(row=r_idx, column=5, value=cell_data.page_number or "")
            ws.cell(row=r_idx, column=6, value=cell_data.confidence)
            ws.cell(row=r_idx, column=7, value=cell_data.matched_phrase or "")
            if not cell_data.value:
                for col in range(1, 8):
                    ws.cell(row=r_idx, column=col).fill = MISSING_FILL
            r_idx += 1

    ws.freeze_panes = f"A{header_row + 1}"
    if ws.max_row > header_row:
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{ws.max_row}"
    _autosize_columns(ws)


def _autosize_columns(ws, max_width: int = 55) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 3, 10), max_width)
